"""
Client routes for Contract Management Platform
"""
import csv
import io
import logging
from datetime import datetime

from flask import (
    Blueprint,
    current_app,
    flash,
    make_response,
    redirect,
    render_template,
    request,
    url_for,
)
from flask_login import current_user, login_required

from app import db
from app.models.client import Client


logger = logging.getLogger(__name__)
clients_bp = Blueprint("clients", __name__)


@clients_bp.route("/")
@login_required
def index():
    """Clients list page with search functionality"""
    try:
        search_term = request.args.get("q", "")
        search_type = request.args.get("type", "")
        status_filter = request.args.get("status", "")
        organization_filter = request.args.get("organization", "")
        page = request.args.get("page", 1, type=int)
        per_page = 20  # Show 20 clients per page

        # Debug logging
        logger.info(f"Client search params - q: '{search_term}', type: '{search_type}', status: '{status_filter}', organization: '{organization_filter}'")
        
        # Build base query
        query = Client.query
        
        # Apply search term filter
        if search_term:
            if search_type == "individual":
                # Search only in name field for individuals
                query = query.filter(Client.name.ilike(f"%{search_term}%"))
            elif search_type == "organization":
                # Search only in organization field
                query = query.filter(Client.organization.ilike(f"%{search_term}%"))
            else:
                # Search in both name, organization, and email (default)
                query = query.filter(
                    db.or_(
                        Client.name.ilike(f"%{search_term}%"),
                        Client.organization.ilike(f"%{search_term}%"),
                        Client.email.ilike(f"%{search_term}%")
                    )
                )
        
        # Apply organization filter
        if organization_filter:
            query = query.filter(Client.organization.ilike(f"%{organization_filter}%"))
        
        # Execute query with pagination
        clients = query.order_by(Client.name).paginate(
            page=page, per_page=per_page, error_out=False
        )

        return render_template(
            "clients/index.html", 
            clients=clients, 
            search_term=search_term,
            search_type=search_type,
            status_filter=status_filter,
            organization_filter=organization_filter
        )

    except Exception as e:
        logger.error(f"Error loading clients: {e}")
        flash("An error occurred while loading clients.", "error")
        # Return empty pagination object on error
        from flask_sqlalchemy import Pagination

        empty_pagination = Pagination(Client.query, 1, 20, 0, [])
        return render_template("clients/index.html", clients=empty_pagination)


@clients_bp.route("/<int:client_id>")
@login_required
def show(client_id):
    """Show client details"""
    try:
        client = Client.query.get_or_404(client_id)

        # Get contracts for this client
        from app.models.contract import Contract

        contracts = (
            Contract.query.filter_by(client_id=client_id, deleted_at=None)
            .order_by(Contract.created_at.desc())
            .all()
        )

        return render_template("clients/show.html", client=client, contracts=contracts)

    except Exception as e:
        logger.error(f"Error showing client {client_id}: {e}")
        flash("An error occurred while loading the client.", "error")
        return redirect(url_for("clients.index"))


@clients_bp.route("/new", methods=["GET", "POST"])
@login_required
def new():
    """Create new client"""
    from app.forms.client_forms import ClientForm

    form = ClientForm()

    if form.validate_on_submit():
        try:
            # Create client from form data
            client = Client(
                name=form.name.data,
                organization=form.organization.data,
                email=form.email.data,
                phone=form.phone.data,
                address=form.address.data,
            )

            # Check for duplicate names
            existing_client = Client.query.filter_by(name=client.name).first()
            if existing_client:
                flash("A client with that name already exists.", "error")
                return render_template("clients/new.html", form=form)

            db.session.add(client)
            db.session.commit()

            current_app.logger.info(
                f"New client created: {client.name} by user {current_user.username}"
            )

            flash(f'Client "{client.name}" created successfully!', "success")
            return redirect(url_for("clients.show", client_id=client.id))

        except Exception as e:
            db.session.rollback()
            logger.error(f"Error creating client: {e}")
            flash("An error occurred while creating the client.", "error")
    elif form.errors:
        # Form validation failed
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text}: {error}", "error")

    # GET request or validation failed - show form
    return render_template("clients/new.html", form=form)


@clients_bp.route("/<int:client_id>/edit", methods=["GET", "POST"])
@login_required
def edit(client_id):
    """Edit client"""
    from app.forms.client_forms import ClientForm

    try:
        client = Client.query.get_or_404(client_id)
        form = ClientForm(obj=client)

        if form.validate_on_submit():
            # Update client data from form
            client.name = form.name.data
            client.organization = form.organization.data
            client.email = form.email.data
            client.phone = form.phone.data
            client.address = form.address.data

            # Check for duplicate names (excluding current client)
            existing_client = Client.query.filter(
                Client.name == client.name, Client.id != client.id
            ).first()
            if existing_client:
                flash("A client with that name already exists.", "error")
                return render_template("clients/edit.html", client=client, form=form)

            db.session.commit()

            current_app.logger.info(
                f"Client updated: {client.name} by user {current_user.username}"
            )

            flash(f'Client "{client.name}" updated successfully!', "success")
            return redirect(url_for("clients.show", client_id=client.id))
        elif form.errors:
            # Form validation failed
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{getattr(form, field).label.text}: {error}", "error")

        # GET request or validation failed - show edit form
        return render_template("clients/edit.html", client=client, form=form)

    except Exception as e:
        logger.error(f"Error editing client {client_id}: {e}")
        flash("An error occurred while editing the client.", "error")
        return redirect(url_for("clients.index"))


@clients_bp.route("/<int:client_id>/delete", methods=["POST"])
@login_required
def delete(client_id):
    """Delete client (only if no contracts exist)"""
    try:
        client = Client.query.get_or_404(client_id)

        # Check if client has ANY contracts (including soft-deleted ones)
        # We cannot delete a client if they have any contracts due to foreign key constraints
        total_contract_count = client.contracts.count()
        active_contract_count = client.contract_count
        
        if total_contract_count > 0:
            if active_contract_count > 0:
                flash(
                    f'Cannot delete client "{client.name}" because they have {active_contract_count} active contract(s). Please delete all contracts first.',
                    "error",
                )
            else:
                flash(
                    f'Cannot delete client "{client.name}" because they have soft-deleted contracts. Please permanently delete all contracts from the admin panel first.',
                    "error",
                )
            return redirect(url_for("clients.show", client_id=client.id))

        client_name = client.name
        db.session.delete(client)
        db.session.commit()

        current_app.logger.info(
            f"Client deleted: {client_name} by user {current_user.username}"
        )

        flash(f'Client "{client_name}" has been deleted.', "success")
        return redirect(url_for("clients.index"))

    except Exception as e:
        db.session.rollback()
        logger.error(f"Error deleting client {client_id}: {e}")
        flash("An error occurred while deleting the client.", "error")
        return redirect(url_for("clients.show", client_id=client_id))


@clients_bp.route("/export", methods=["GET", "POST"])
@login_required
def export_clients():
    """Export clients with current filters applied"""
    try:
        # Get filter parameters from request
        search_term = request.form.get("q", "") or request.args.get("q", "")
        search_type = request.form.get("type", "") or request.args.get("type", "")
        status_filter = request.form.get("status", "") or request.args.get("status", "")
        organization_filter = request.form.get("organization", "") or request.args.get("organization", "")
        export_format = request.form.get("export_format", "csv")
        
        logger.info(f"Exporting clients - format: {export_format}, filters: q='{search_term}', type='{search_type}', org='{organization_filter}'")
        
        # Build query with same logic as index route
        query = Client.query
        
        # Apply search term filter
        if search_term:
            if search_type == "individual":
                # Search only in name field for individuals
                query = query.filter(Client.name.ilike(f"%{search_term}%"))
            elif search_type == "organization":
                # Search only in organization field
                query = query.filter(Client.organization.ilike(f"%{search_term}%"))
            else:
                # Search in both name, organization, and email (default)
                query = query.filter(
                    db.or_(
                        Client.name.ilike(f"%{search_term}%"),
                        Client.organization.ilike(f"%{search_term}%"),
                        Client.email.ilike(f"%{search_term}%")
                    )
                )
        
        # Apply organization filter
        if organization_filter:
            query = query.filter(Client.organization.ilike(f"%{organization_filter}%"))
        
        # Get all matching clients (no pagination for export)
        clients = query.order_by(Client.name).all()
        
        if export_format.lower() == "csv":
            return export_clients_csv(clients)
        elif export_format.lower() == "excel":
            return export_clients_excel(clients)
        elif export_format.lower() == "pdf":
            return export_clients_pdf(clients)
        else:
            flash("Unsupported export format.", "error")
            return redirect(url_for("clients.index"))
            
    except Exception as e:
        logger.error(f"Error exporting clients: {e}")
        flash("An error occurred while exporting clients.", "error")
        return redirect(url_for("clients.index"))


def export_clients_csv(clients):
    """Export clients to CSV format"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write CSV headers
        headers = [
            'ID', 'Name', 'Organization', 'Email', 'Phone', 'Address', 
            'Contract Count', 'Created Date', 'Updated Date'
        ]
        writer.writerow(headers)
        
        # Write client data
        for client in clients:
            row = [
                client.id,
                client.name,
                client.organization or '',
                client.email or '',
                client.phone or '',
                client.address or '',
                client.contract_count,
                client.created_at.strftime('%Y-%m-%d') if client.created_at else '',
                client.updated_at.strftime('%Y-%m-%d') if client.updated_at else ''
            ]
            writer.writerow(row)
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=clients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        logger.info(f"Exported {len(clients)} clients to CSV")
        return response
        
    except Exception as e:
        logger.error(f"Error creating CSV export: {e}")
        raise


def export_clients_excel(clients):
    """Export clients to Excel format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Clients Export"
        
        # Define headers
        headers = [
            'ID', 'Name', 'Organization', 'Email', 'Phone', 'Address', 
            'Contract Count', 'Created Date', 'Updated Date'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write and style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write client data
        for row_num, client in enumerate(clients, 2):
            data = [
                client.id,
                client.name,
                client.organization or '',
                client.email or '',
                client.phone or '',
                client.address or '',
                client.contract_count,
                client.created_at.strftime('%Y-%m-%d') if client.created_at else '',
                client.updated_at.strftime('%Y-%m-%d') if client.updated_at else ''
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(headers[col-1])
            for row in range(2, len(clients) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 40)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=clients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        logger.info(f"Exported {len(clients)} clients to Excel")
        return response
        
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise


def export_clients_pdf(clients):
    """Export clients to PDF format"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Create PDF buffer
        buffer = io.BytesIO()
        
        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        )
        
        # Build content
        content = []
        
        # Add title
        title = Paragraph("Clients Export Report", title_style)
        content.append(title)
        
        # Add export info
        export_info = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>Total Clients: {len(clients)}"
        info_para = Paragraph(export_info, styles['Normal'])
        content.append(info_para)
        content.append(Spacer(1, 20))
        
        if clients:
            # Prepare table data
            table_data = [
                ['ID', 'Name', 'Organization', 'Email', 'Phone', 'Contracts', 'Created Date']
            ]
            
            for client in clients:
                row = [
                    str(client.id),
                    client.name[:25] + ('...' if len(client.name) > 25 else ''),
                    (client.organization[:20] + ('...' if len(client.organization) > 20 else '')) if client.organization else 'Individual',
                    client.email[:30] + ('...' if len(client.email or '') > 30 else '') if client.email else 'N/A',
                    client.phone or 'N/A',
                    str(client.contract_count),
                    client.created_at.strftime('%Y-%m-%d') if client.created_at else 'N/A'
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1.3*inch, 1.8*inch, 1*inch, 0.8*inch, 1*inch])
            
            # Apply table style
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (5, 1), (5, -1), 'CENTER'),  # Contract count column center
                
                # Grid and borders
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            content.append(table)
        else:
            no_data = Paragraph("No clients found matching the current filters.", styles['Normal'])
            content.append(no_data)
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=clients_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        logger.info(f"Exported {len(clients)} clients to PDF")
        return response
        
    except Exception as e:
        logger.error(f"Error creating PDF export: {e}")
        raise





@clients_bp.route("/api/search")
@login_required
def api_search():
    """API endpoint for client search"""
    try:
        search_term = request.args.get("q", "")

        if search_term:
            clients = (
                Client.query.filter(
                    db.or_(
                        Client.name.ilike(f"%{search_term}%"),
                        Client.organization.ilike(f"%{search_term}%"),
                    )
                )
                .limit(10)
                .all()
            )
        else:
            clients = []

        # Convert to JSON-serializable format
        results = []
        for client in clients:
            results.append(client.to_dict())

        return {"clients": results}

    except Exception as e:
        logger.error(f"Error in API client search: {e}")
        return {"error": "Search failed"}, 500

