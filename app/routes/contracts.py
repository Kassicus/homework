"""
Contract routes for Contract Management Platform
"""
import csv
import io
import logging
import os
from datetime import datetime

from flask import (
    Blueprint,
    current_app,
    flash,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    url_for,
)
from flask_login import current_user, login_required

from app import db
from app.models.client import Client
from app.models.contract import Contract
from app.models.contract_document import ContractDocument
from app.services.contract_service import ContractService
from app.services.docusign_service import docusign_service
from app.services.document_service import DocumentService
from app.services.file_service import FileService
from app.utils.activity_decorators import (
    log_view, log_create, log_update, log_delete, log_restore,
    log_docusign_send, log_docusign_check, log_docusign_void,
    log_document_upload, log_document_delete, log_document_download, log_document_set_primary,
    get_contract_info, get_docusign_info, get_document_info
)
from app.models.activity_log import ActivityLog


logger = logging.getLogger(__name__)
contracts_bp = Blueprint("contracts", __name__)


@contracts_bp.route("/")
@login_required
def index():
    """Contracts list page with search functionality"""
    search_term = request.args.get("q", "")
    status_filter = request.args.get("status", "")
    client_filter = request.args.get("client", "")
    contract_type_filter = request.args.get("contract_type", "")
    docusign_status_filter = request.args.get("docusign_status", "")
    date_from = request.args.get("date_from", "")
    date_to = request.args.get("date_to", "")
    expiring_soon = request.args.get("expiring_soon", "")
    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    try:
        # Debug logging
        logger.info(f"Contract search params - q: '{search_term}', status: '{status_filter}', client: '{client_filter}', type: '{contract_type_filter}', docusign: '{docusign_status_filter}', date_from: '{date_from}', date_to: '{date_to}', expiring_soon: '{expiring_soon}'")
        
        if search_term or status_filter or client_filter or contract_type_filter or docusign_status_filter or date_from or date_to or expiring_soon:
            # Build filters for search
            filters = {}
            if status_filter:
                filters["status"] = status_filter
            if client_filter:
                filters["client_id"] = int(client_filter)
            if contract_type_filter:
                filters["contract_type"] = contract_type_filter
            if date_from:
                filters["date_from"] = date_from
            if date_to:
                filters["date_to"] = date_to
            if expiring_soon:
                filters["expiring_soon"] = int(expiring_soon)

            # Perform search
            contracts = ContractService.search_contracts(
                search_term=search_term, filters=filters, page=page, per_page=per_page
            )
        else:
            # Show all contracts when no search parameters
            contracts = ContractService.get_all_contracts(
                include_deleted=False, page=page, per_page=per_page
            )

        # Get clients for filter dropdown
        clients = Client.query.all()
        
        # Get contract types for filter dropdown
        contract_types = ContractService.get_contract_type_choices_for_search()

        return render_template(
            "contracts/index.html",
            contracts=contracts,
            search_term=search_term,
            filters=request.args,
            clients=clients,
            contract_types=contract_types,
        )

    except Exception as e:
        logger.error(f"Error loading contracts: {e}")
        flash("An error occurred while loading contracts.", "error")
        return render_template("contracts/index.html", contracts=None)





@contracts_bp.route("/<int:contract_id>")
@login_required
@log_view(ActivityLog.RESOURCE_CONTRACT, get_contract_info)
def show(contract_id):
    """Show contract details"""
    try:
        # First, try to get the contract directly from the database
        contract = Contract.query.get(contract_id)

        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))

        # Check if contract is soft-deleted
        if contract.deleted_at:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))

        # Log access with request data (optional)
        try:
            # Log access - table existence is handled by the model
            contract.log_access(
                user_id=current_user.id,
                access_type="view",
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string,
            )
            db.session.commit()
        except Exception as access_error:
            # Log access error but don't fail the request
            current_app.logger.warning(f"Failed to log contract access: {access_error}")
            db.session.rollback()

        return render_template("contracts/show.html", contract=contract)

    except Exception as e:
        logger.error(f"Error showing contract {contract_id}: {e}")
        # Log the full traceback for debugging
        import traceback

        logger.error(f"Full traceback: {traceback.format_exc()}")
        flash("An error occurred while loading the contract.", "error")
        return redirect(url_for("contracts.index"))


@contracts_bp.route("/new", methods=["GET", "POST"])
@login_required
@log_create(ActivityLog.RESOURCE_CONTRACT, lambda *args, **kwargs: (None, None))
def new():
    """Create new contract"""
    from app.forms.contract_forms import ContractForm

    # Get clients first
    clients = Client.query.order_by(Client.name).all()

    # Create form with client choices
    form = ContractForm()
    form.client_id.choices = [(0, "Select a client...")] + [
        (c.id, c.name) for c in clients
    ]
    
    # Set contract type choices dynamically
    form.contract_type.choices = ContractService.get_contract_type_choices_for_forms()

    # Set default values for other select fields
    if not form.contract_type.data:
        form.contract_type.data = ""
    if not form.status.data:
        form.status.data = ""

    if form.validate_on_submit():
        # Additional validation for client_id
        if form.client_id.data == 0:
            flash("Please select a client.", "error")
            return render_template("contracts/new.html", form=form, clients=clients)

        try:
            # Handle custom contract type
            contract_type = form.contract_type.data
            if contract_type == "custom" and form.custom_contract_type.data:
                contract_type = form.custom_contract_type.data.strip()
                # Validate and potentially create the new type
                if ContractService.create_contract_type_if_not_exists(contract_type):
                    logger.info(f"Using custom contract type: {contract_type}")
                else:
                    flash("Invalid custom contract type.", "error")
                    return render_template("contracts/new.html", form=form, clients=clients)
            elif contract_type == "custom":
                flash("Please enter a custom contract type or select an existing one.", "error")
                return render_template("contracts/new.html", form=form, clients=clients)

            # Get form data
            contract_data = {
                "title": form.title.data,
                "description": form.description.data,
                "client_id": form.client_id.data,
                "contract_type": contract_type,
                "status": form.status.data,
                "contract_value": form.contract_value.data,
                "effective_date": form.effective_date.data,
                "expiration_date": form.expiration_date.data,
                "renewal_date": form.renewal_date.data,
            }

            # Get uploaded file
            file = form.contract_file.data

            # Create contract
            contract = ContractService.create_contract(
                contract_data=contract_data, file=file, created_by=current_user.id
            )

            flash(f'Contract "{contract.title}" created successfully!', "success")
            return redirect(url_for("contracts.show", contract_id=contract.id))

        except Exception as e:
            logger.error(f"Error creating contract: {e}")
            flash(f"An error occurred while creating the contract: {str(e)}", "error")
    elif form.errors:
        # Form validation failed
        for field, errors in form.errors.items():
            for error in errors:
                flash(f"{getattr(form, field).label.text}: {error}", "error")

    # GET request or validation failed - show form
    return render_template("contracts/new.html", form=form, clients=clients)


@contracts_bp.route("/<int:contract_id>/edit", methods=["GET", "POST"])
@login_required
@log_update(ActivityLog.RESOURCE_CONTRACT, get_contract_info)
def edit(contract_id):
    """Edit contract"""
    from app.forms.contract_forms import ContractForm

    try:
        contract = ContractService.get_contract_by_id(contract_id)

        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))

        # Get clients first
        clients = Client.query.order_by(Client.name).all()

        # Create form with client choices
        form = ContractForm(obj=contract)
        form.client_id.choices = [(0, "Select a client...")] + [
            (c.id, c.name) for c in clients
        ]
        
        # Set contract type choices dynamically
        form.contract_type.choices = ContractService.get_contract_type_choices_for_forms()

        # Set default values for other select fields if they're None
        if not form.contract_type.data:
            form.contract_type.data = ""
        if not form.status.data:
            form.status.data = ""

        if form.validate_on_submit():
            # Additional validation for client_id
            if form.client_id.data == 0:
                flash("Please select a client.", "error")
                return render_template(
                    "contracts/edit.html", contract=contract, form=form, clients=clients
                )

            # Handle custom contract type
            contract_type = form.contract_type.data
            if contract_type == "custom" and form.custom_contract_type.data:
                contract_type = form.custom_contract_type.data.strip()
                # Validate and potentially create the new type
                if ContractService.create_contract_type_if_not_exists(contract_type):
                    logger.info(f"Using custom contract type: {contract_type}")
                else:
                    flash("Invalid custom contract type.", "error")
                    return render_template(
                        "contracts/edit.html", contract=contract, form=form, clients=clients
                    )
            elif contract_type == "custom":
                flash("Please enter a custom contract type or select an existing one.", "error")
                return render_template(
                    "contracts/edit.html", contract=contract, form=form, clients=clients
                )

            # Get form data
            update_data = {
                "title": form.title.data,
                "description": form.description.data,
                "client_id": form.client_id.data,
                "contract_type": contract_type,
                "status": form.status.data,
                "contract_value": form.contract_value.data,
                "effective_date": form.effective_date.data,
                "expiration_date": form.expiration_date.data,
                "renewal_date": form.renewal_date.data,
            }

            # Get uploaded file if provided
            file = form.contract_file.data if form.contract_file.data else None

            # Update contract
            contract = ContractService.update_contract(
                contract_id=contract_id,
                update_data=update_data,
                updated_by=current_user.id,
                file=file,
            )

            flash(f'Contract "{contract.title}" updated successfully!', "success")
            return redirect(url_for("contracts.show", contract_id=contract.id))
        elif form.errors:
            # Form validation failed
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{getattr(form, field).label.text}: {error}", "error")

        # GET request or validation failed - show edit form
        return render_template(
            "contracts/edit.html", contract=contract, form=form, clients=clients
        )

    except Exception as e:
        logger.error(f"Error editing contract {contract_id}: {e}")
        flash("An error occurred while editing the contract.", "error")
        return redirect(url_for("contracts.index"))


@contracts_bp.route("/<int:contract_id>/delete", methods=["POST"])
@login_required
@log_delete(ActivityLog.RESOURCE_CONTRACT, get_contract_info)
def delete(contract_id):
    """Soft delete contract"""
    try:
        contract = ContractService.soft_delete_contract(contract_id, current_user)

        flash(f'Contract "{contract.title}" has been deleted.', "success")
        return redirect(url_for("contracts.index"))

    except Exception as e:
        logger.error(f"Error deleting contract {contract_id}: {e}")
        flash("An error occurred while deleting the contract.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/restore", methods=["POST"])
@login_required
@log_restore(ActivityLog.RESOURCE_CONTRACT, get_contract_info)
def restore(contract_id):
    """Restore soft-deleted contract"""
    try:
        contract = ContractService.restore_contract(contract_id, current_user)

        flash(f'Contract "{contract.title}" has been restored.', "success")
        return redirect(url_for("contracts.show", contract_id=contract.id))

    except Exception as e:
        logger.error(f"Error restoring contract {contract_id}: {e}")
        flash("An error occurred while restoring the contract.", "error")
        return redirect(url_for("contracts.index"))


@contracts_bp.route("/<int:contract_id>/status", methods=["POST"])
@login_required
def update_status(contract_id):
    """Update contract status"""
    try:
        new_status = request.form.get("status")
        reason = request.form.get("reason", "")

        if not new_status:
            flash("Status is required.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))

        contract = ContractService.update_contract_status(
            contract_id=contract_id,
            new_status=new_status,
            changed_by=current_user,
            reason=reason,
        )

        flash(f"Contract status updated to {new_status}.", "success")
        return redirect(url_for("contracts.show", contract_id=contract.id))

    except Exception as e:
        logger.error(f"Error updating contract status {contract_id}: {e}")
        flash("An error occurred while updating the contract status.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/documents/upload", methods=["GET", "POST"])
@login_required
@log_document_upload(get_document_info)
def upload_document(contract_id):
    """Upload document to existing contract (supports multiple documents)"""
    from app.forms.contract_forms import DocumentUploadForm
    
    try:
        contract = ContractService.get_contract_by_id(contract_id)
        
        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))
        
        # Migrate legacy document if needed
        DocumentService.migrate_legacy_document(contract)
        
        form = DocumentUploadForm()
        
        if form.validate_on_submit():
            try:
                # Get uploaded file and form data
                file = form.contract_file.data
                document_type = form.document_type.data
                description = form.description.data if form.description.data else None
                is_primary = form.is_primary.data == "true"
                
                # Upload document using new service
                document = DocumentService.upload_document(
                    contract_id=contract_id,
                    file=file,
                    uploaded_by=current_user.id,
                    document_type=document_type,
                    description=description,
                    is_primary=is_primary
                )
                
                flash(f'Document "{document.original_filename}" uploaded successfully!', "success")
                return redirect(url_for("contracts.show", contract_id=contract.id))
                
            except Exception as file_error:
                logger.error(f"Error uploading document to contract {contract_id}: {file_error}")
                flash(f"An error occurred while uploading the document: {str(file_error)}", "error")
        
        elif form.errors:
            # Form validation failed
            for field, errors in form.errors.items():
                for error in errors:
                    flash(f"{getattr(form, field).label.text}: {error}", "error")
        
        # Get existing documents for display
        existing_documents = DocumentService.get_contract_documents(contract_id)
        
        # Show upload form (GET request or validation failed)
        return render_template("contracts/upload.html", contract=contract, form=form, existing_documents=existing_documents)
        
    except Exception as e:
        logger.error(f"Error in document upload for contract {contract_id}: {e}")
        flash("An error occurred while processing the upload.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/documents/<int:document_id>/delete", methods=["POST"])
@login_required
@log_document_delete(get_document_info)
def delete_document(contract_id, document_id):
    """Delete a specific document from a contract"""
    try:
        contract = ContractService.get_contract_by_id(contract_id)
        
        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))
        
        document = DocumentService.get_document_by_id(document_id)
        
        if not document or document.contract_id != contract_id:
            flash("Document not found.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Store filename for success message
        filename = document.original_filename
        
        # Delete document
        DocumentService.delete_document(document_id, contract_id)
        
        flash(f'Document "{filename}" has been deleted.', "success")
        return redirect(url_for("contracts.show", contract_id=contract_id))
        
    except Exception as e:
        logger.error(f"Error deleting document {document_id} from contract {contract_id}: {e}")
        flash("An error occurred while deleting the document.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/documents/<int:document_id>/set-primary", methods=["POST"])
@login_required
@log_document_set_primary(get_document_info)
def set_primary_document(contract_id, document_id):
    """Set a document as the primary document for a contract"""
    try:
        contract = ContractService.get_contract_by_id(contract_id)
        
        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))
        
        document = DocumentService.set_primary_document(document_id, contract_id)
        
        flash(f'"{document.original_filename}" is now the primary document.', "success")
        return redirect(url_for("contracts.show", contract_id=contract_id))
        
    except Exception as e:
        logger.error(f"Error setting primary document {document_id} for contract {contract_id}: {e}")
        flash("An error occurred while setting the primary document.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/documents/<int:document_id>/download")
@login_required
@log_document_download(get_document_info)
def download_document(contract_id, document_id):
    """Download a specific document"""
    try:
        contract = ContractService.get_contract_by_id(contract_id)
        
        if not contract:
            flash("Contract not found.", "error")
            return redirect(url_for("contracts.index"))
        
        document = DocumentService.get_document_by_id(document_id)
        
        if not document or document.contract_id != contract_id:
            flash("Document not found.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Security and file existence checks
        if not os.path.exists(document.file_path):
            flash("Document file not found on server.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Log access
        try:
            contract.log_access(
                user_id=current_user.id,
                access_type="download",
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string,
            )
            db.session.commit()
        except Exception as access_error:
            logger.warning(f"Failed to log document download access: {access_error}")
            db.session.rollback()
        
        # Send file
        from flask import send_file
        return send_file(
            document.file_path,
            as_attachment=True,
            download_name=document.original_filename,
            mimetype=document.mime_type
        )
        
    except Exception as e:
        logger.error(f"Error downloading document {document_id} from contract {contract_id}: {e}")
        flash("An error occurred while downloading the document.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/download")
@login_required
def download(contract_id):
    """Download contract file"""
    try:
        contract = ContractService.get_contract_by_id(contract_id)

        if not contract or not contract.file_path:
            flash("Contract file not found.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))

        # Security check: Ensure file path is within upload folder
        import os

        from flask import abort, current_app, send_file

        upload_folder = current_app.config.get("UPLOAD_FOLDER", "uploads")
        file_path = contract.file_path

        # Convert relative path to absolute path
        if not os.path.isabs(file_path):
            # If the path starts with 'uploads/', remove it to avoid duplication
            if file_path.startswith("uploads/"):
                file_path = file_path[8:]  # Remove 'uploads/' prefix
            file_path = os.path.join(upload_folder, file_path)

        # Debug logging
        logger.info(f"Original file_path from DB: {contract.file_path}")
        logger.info(f"Upload folder: {upload_folder}")
        logger.info(f"Resolved file_path: {file_path}")
        logger.info(f"File exists: {os.path.exists(file_path)}")

        # Normalize paths to prevent directory traversal attacks
        abs_upload_folder = os.path.abspath(upload_folder)
        abs_file_path = os.path.abspath(file_path)

        if not abs_file_path.startswith(abs_upload_folder):
            logger.error(
                f"Security violation: Attempted to access file outside upload folder: {file_path}"
            )
            abort(403, description="Access denied")

        # Check if file exists
        if not os.path.exists(file_path):
            logger.error(f"File not found on server: {file_path}")
            flash(
                "File not found on server. The file may have been moved or deleted.",
                "error",
            )
            return redirect(url_for("contracts.show", contract_id=contract_id))

        # Check if file is readable
        if not os.access(file_path, os.R_OK):
            logger.error(f"File not readable: {file_path}")
            flash("File access denied. Please contact an administrator.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))

        # Log download access (optional - don't fail download if logging fails)
        try:
            contract.log_access(
                user_id=current_user.id,
                access_type="download",
                ip_address=request.remote_addr,
                user_agent=request.user_agent.string,
            )
            db.session.commit()
        except Exception as access_error:
            logger.warning(f"Failed to log download access: {access_error}")
            # Continue with download even if logging fails
            db.session.rollback()

        # Determine MIME type for proper browser handling
        mime_type = contract.mime_type or "application/octet-stream"

        # Set filename for download (use original filename if available)
        filename = contract.file_name or os.path.basename(file_path)

        # Send file with proper headers
        return send_file(
            file_path, as_attachment=True, download_name=filename, mimetype=mime_type
        )

    except Exception as e:
        logger.error(f"Error downloading contract {contract_id}: {e}")
        flash("An error occurred while downloading the contract.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/export", methods=["GET", "POST"])
@login_required
def export_contracts():
    """Export contracts with current filters applied"""
    try:
        # Get filter parameters from request
        search_term = request.form.get("q", "") or request.args.get("q", "")
        status_filter = request.form.get("status", "") or request.args.get("status", "")
        client_filter = request.form.get("client", "") or request.args.get("client", "")
        contract_type_filter = request.form.get("contract_type", "") or request.args.get("contract_type", "")
        docusign_status_filter = request.form.get("docusign_status", "") or request.args.get("docusign_status", "")
        date_from = request.form.get("date_from", "") or request.args.get("date_from", "")
        date_to = request.form.get("date_to", "") or request.args.get("date_to", "")
        expiring_soon = request.form.get("expiring_soon", "") or request.args.get("expiring_soon", "")
        export_format = request.form.get("export_format", "csv")
        
        logger.info(f"Exporting contracts - format: {export_format}, filters: q='{search_term}', status='{status_filter}', client='{client_filter}', type='{contract_type_filter}', docusign='{docusign_status_filter}'")
        
        # Build filters for export (same logic as search)
        if search_term or status_filter or client_filter or contract_type_filter or docusign_status_filter or date_from or date_to or expiring_soon:
            filters = {}
            if status_filter:
                filters["status"] = status_filter
            if client_filter:
                filters["client_id"] = int(client_filter)
            if contract_type_filter:
                filters["contract_type"] = contract_type_filter
            if date_from:
                filters["date_from"] = date_from
            if date_to:
                filters["date_to"] = date_to
            if expiring_soon:
                filters["expiring_soon"] = int(expiring_soon)
                
            # Get filtered contracts (all pages, no pagination for export)
            contracts_paginated = ContractService.search_contracts(
                search_term=search_term, filters=filters, page=1, per_page=10000
            )
            contracts = contracts_paginated.items
        else:
            # Export all contracts if no filters
            contracts_paginated = ContractService.get_all_contracts(
                include_deleted=False, page=1, per_page=10000
            )
            contracts = contracts_paginated.items
        
        if export_format.lower() == "csv":
            return export_contracts_csv(contracts)
        elif export_format.lower() == "excel":
            return export_contracts_excel(contracts)
        elif export_format.lower() == "pdf":
            return export_contracts_pdf(contracts)
        else:
            flash("Unsupported export format.", "error")
            return redirect(url_for("contracts.index"))
            
    except Exception as e:
        logger.error(f"Error exporting contracts: {e}")
        flash("An error occurred while exporting contracts.", "error")
        return redirect(url_for("contracts.index"))


def export_contracts_csv(contracts):
    """Export contracts to CSV format"""
    try:
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write CSV headers
        headers = [
            'ID', 'Title', 'Description', 'Client', 'Contract Type', 
            'Status', 'Contract Value', 'Created Date', 'Effective Date', 
            'Expiration Date', 'Renewal Date', 'Created By', 'File Name'
        ]
        writer.writerow(headers)
        
        # Write contract data
        for contract in contracts:
            row = [
                contract.id,
                contract.title,
                contract.description or '',
                contract.client.name if contract.client else '',
                contract.contract_type or '',
                contract.status,
                contract.docusign_status or 'Not Sent',
                float(contract.contract_value) if contract.contract_value else '',
                contract.created_date.strftime('%Y-%m-%d') if contract.created_date else '',
                contract.effective_date.strftime('%Y-%m-%d') if contract.effective_date else '',
                contract.expiration_date.strftime('%Y-%m-%d') if contract.expiration_date else '',
                contract.renewal_date.strftime('%Y-%m-%d') if contract.renewal_date else '',
                contract.creator.username if contract.creator else '',
                contract.legacy_file_name or ''
            ]
            writer.writerow(row)
        
        # Create response
        output.seek(0)
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'text/csv'
        response.headers['Content-Disposition'] = f'attachment; filename=contracts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.csv'
        
        logger.info(f"Exported {len(contracts)} contracts to CSV")
        return response
        
    except Exception as e:
        logger.error(f"Error creating CSV export: {e}")
        raise


def export_contracts_excel(contracts):
    """Export contracts to Excel format"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create workbook and worksheet
        wb = Workbook()
        ws = wb.active
        ws.title = "Contracts Export"
        
        # Define headers
        headers = [
            'ID', 'Title', 'Description', 'Client', 'Contract Type', 
            'Status', 'DocuSign Status', 'Contract Value', 'Created Date', 'Effective Date', 
            'Expiration Date', 'Renewal Date', 'Created By', 'File Name'
        ]
        
        # Style for headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Write and style headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Write contract data
        for row_num, contract in enumerate(contracts, 2):
            data = [
                contract.id,
                contract.title,
                contract.description or '',
                contract.client.name if contract.client else '',
                contract.contract_type or '',
                contract.status,
                contract.docusign_status or 'Not Sent',
                float(contract.contract_value) if contract.contract_value else '',
                contract.created_date.strftime('%Y-%m-%d') if contract.created_date else '',
                contract.effective_date.strftime('%Y-%m-%d') if contract.effective_date else '',
                contract.expiration_date.strftime('%Y-%m-%d') if contract.expiration_date else '',
                contract.renewal_date.strftime('%Y-%m-%d') if contract.renewal_date else '',
                contract.creator.username if contract.creator else '',
                contract.legacy_file_name or ''
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row_num, column=col, value=value)
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            column_letter = get_column_letter(col)
            max_length = len(headers[col-1])
            for row in range(2, len(contracts) + 2):
                cell_value = str(ws.cell(row=row, column=col).value or '')
                max_length = max(max_length, len(cell_value))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        
        # Save to BytesIO
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename=contracts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        
        logger.info(f"Exported {len(contracts)} contracts to Excel")
        return response
        
    except Exception as e:
        logger.error(f"Error creating Excel export: {e}")
        raise


def export_contracts_pdf(contracts):
    """Export contracts to PDF format"""
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.enums import TA_CENTER, TA_LEFT
        
        # Create PDF buffer
        buffer = io.BytesIO()
        
        # Create document
        doc = SimpleDocTemplate(
            buffer,
            pagesize=A4,
            rightMargin=72,
            leftMargin=72,
            topMargin=72,
            bottomMargin=18
        )
        
        # Get styles
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            spaceAfter=30,
            alignment=TA_CENTER,
            textColor=colors.HexColor('#2c3e50')
        )
        
        # Build content
        content = []
        
        # Add title
        title = Paragraph("Contracts Export Report", title_style)
        content.append(title)
        
        # Add export info
        export_info = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}<br/>Total Contracts: {len(contracts)}"
        info_para = Paragraph(export_info, styles['Normal'])
        content.append(info_para)
        content.append(Spacer(1, 20))
        
        if contracts:
            # Prepare table data
            table_data = [
                ['ID', 'Title', 'Client', 'Type', 'Status', 'DocuSign', 'Value', 'Effective', 'Expires']
            ]
            
            for contract in contracts:
                row = [
                    str(contract.id),
                    contract.title[:30] + ('...' if len(contract.title) > 30 else ''),
                    (contract.client.name[:20] + ('...' if len(contract.client.name) > 20 else '')) if contract.client else 'N/A',
                    contract.contract_type[:15] + ('...' if len(contract.contract_type or '') > 15 else '') if contract.contract_type else 'N/A',
                    contract.status,
                    contract.docusign_status[:10] + ('...' if len(contract.docusign_status or '') > 10 else '') if contract.docusign_status else 'Not Sent',
                    f'${contract.contract_value:,.2f}' if contract.contract_value else 'N/A',
                    contract.effective_date.strftime('%Y-%m-%d') if contract.effective_date else 'N/A',
                    contract.expiration_date.strftime('%Y-%m-%d') if contract.expiration_date else 'N/A'
                ]
                table_data.append(row)
            
            # Create table
            table = Table(table_data, colWidths=[0.5*inch, 1.4*inch, 1*inch, 0.8*inch, 0.7*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch])
            
            # Apply table style
            table.setStyle(TableStyle([
                # Header row styling
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 10),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                
                # Data rows styling
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 8),
                ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                ('ALIGN', (4, 1), (4, -1), 'CENTER'),  # Status column center
                ('ALIGN', (5, 1), (5, -1), 'RIGHT'),   # Value column right
                
                # Grid and borders
                ('GRID', (0, 0), (-1, -1), 1, colors.black),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                
                # Alternating row colors
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')])
            ]))
            
            content.append(table)
        else:
            no_data = Paragraph("No contracts found matching the current filters.", styles['Normal'])
            content.append(no_data)
        
        # Build PDF
        doc.build(content)
        buffer.seek(0)
        
        # Create response
        response = make_response(buffer.getvalue())
        response.headers['Content-Type'] = 'application/pdf'
        response.headers['Content-Disposition'] = f'attachment; filename=contracts_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
        
        logger.info(f"Exported {len(contracts)} contracts to PDF")
        return response
        
    except Exception as e:
        logger.error(f"Error creating PDF export: {e}")
        raise


@contracts_bp.route("/deleted")
@login_required
def deleted():
    """Show deleted contracts"""
    if not current_user.is_admin:
        flash("Access denied. Admin privileges required.", "error")
        return redirect(url_for("contracts.index"))

    page = request.args.get("page", 1, type=int)
    per_page = request.args.get("per_page", 20, type=int)

    try:
        contracts = ContractService.get_all_contracts(
            include_deleted=True, page=page, per_page=per_page
        )

        return render_template("contracts/deleted.html", contracts=contracts)

    except Exception as e:
        logger.error(f"Error loading deleted contracts: {e}")
        flash("An error occurred while loading deleted contracts.", "error")
        return render_template("contracts/deleted.html", contracts=None)


@contracts_bp.route("/api/search")
@login_required
def api_search():
    """API endpoint for contract search"""
    try:
        search_term = request.args.get("q", "")
        page = request.args.get("page", 1, type=int)
        per_page = request.args.get("per_page", 10, type=int)

        contracts = ContractService.search_contracts(
            search_term=search_term, page=page, per_page=per_page
        )

        # Convert to JSON-serializable format
        results = []
        for contract in contracts.items:
            results.append(contract.to_dict())

        return jsonify(
            {
                "contracts": results,
                "total": contracts.total,
                "pages": contracts.pages,
                "current_page": contracts.page,
                "has_next": contracts.has_next,
                "has_prev": contracts.has_prev,
            }
        )

    except Exception as e:
        logger.error(f"Error in API search: {e}")
        return jsonify({"error": "Search failed"}), 500


@contracts_bp.route("/<int:contract_id>/send_for_signature", methods=["POST"])
@login_required
@log_docusign_send(get_docusign_info)
def send_for_signature(contract_id):
    """Send contract to client via DocuSign (Mock Implementation)"""
    try:
        contract = Contract.query.get_or_404(contract_id)
        
        # Check if contract has documents
        if not contract.documents.first() and not contract.file_path:
            flash("Cannot send contract without documents. Please upload a document first.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Get recipient details and document selection from form
        recipient_email = request.form.get('recipient_email', '').strip()
        recipient_name = request.form.get('recipient_name', '').strip()
        document_id = request.form.get('document_id', '').strip()
        
        # Validate email
        if not recipient_email:
            flash("Recipient email is required.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Validate document selection
        if not document_id:
            flash("Please select a document to send for signature.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Verify selected document belongs to this contract
        from app.models.contract_document import ContractDocument
        selected_document = ContractDocument.query.filter_by(
            id=document_id, 
            contract_id=contract_id
        ).first()
        
        if not selected_document:
            flash("Invalid document selection.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Default recipient name to client name if not provided
        if not recipient_name and contract.client:
            recipient_name = contract.client.name
        
        # Send to DocuSign (mock)
        result = docusign_service.send_contract_for_signature(
            contract_id, recipient_email, recipient_name, document_id
        )
        
        if result['success']:
            flash(f"✅ Contract sent to {recipient_email} for signature! (Mock Mode)", "success")
            logger.info(f"Contract {contract_id} sent for signature to {recipient_email}")
        else:
            flash(f"❌ Error sending contract: {result['error']}", "error")
            logger.error(f"Failed to send contract {contract_id}: {result['error']}")
            
        return redirect(url_for("contracts.show", contract_id=contract_id))
        
    except Exception as e:
        logger.error(f"Error sending contract for signature: {e}")
        flash("An error occurred while sending the contract.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/<int:contract_id>/docusign_status", methods=["GET"])
@login_required
@log_docusign_check(get_docusign_info)
def check_docusign_status(contract_id):
    """Check DocuSign status for a contract (Mock Implementation)"""
    try:
        contract = Contract.query.get_or_404(contract_id)
        
        if not contract.docusign_envelope_id:
            return jsonify({
                'success': False,
                'error': 'No DocuSign envelope found for this contract'
            }), 404
        
        # Check status (mock)
        result = docusign_service.check_envelope_status(
            contract.docusign_envelope_id, contract_id
        )
        
        return jsonify(result)
        
    except Exception as e:
        logger.error(f"Error checking DocuSign status: {e}")
        return jsonify({
            'success': False,
            'error': str(e)
        }), 500


@contracts_bp.route("/<int:contract_id>/void_docusign", methods=["POST"])
@login_required
@log_docusign_void(get_docusign_info)
def void_docusign_envelope(contract_id):
    """Void a DocuSign envelope (Mock Implementation)"""
    try:
        contract = Contract.query.get_or_404(contract_id)
        
        if not contract.docusign_envelope_id:
            flash("No DocuSign envelope found to void.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        if contract.docusign_status == 'completed':
            flash("Cannot void a completed envelope.", "error")
            return redirect(url_for("contracts.show", contract_id=contract_id))
        
        # Store old status for logging
        old_status = contract.docusign_status
        envelope_id = contract.docusign_envelope_id
        
        # Mock voiding
        contract.docusign_status = 'voided'
        db.session.commit()
        
        # Log additional details about the void operation
        try:
            from app.services.activity_service import log_user_activity
            log_user_activity(
                action=ActivityLog.ACTION_DOCUSIGN_STATUS_CHANGE,
                resource_type=ActivityLog.RESOURCE_CONTRACT,
                resource_id=contract_id,
                resource_title=f"{contract.title} (Envelope: {envelope_id[:8]}...)",
                success=True,
                additional_data={
                    'old_status': old_status,
                    'new_status': 'voided',
                    'envelope_id': envelope_id,
                    'action_type': 'manual_void',
                    'mock_service': True
                }
            )
        except Exception as log_error:
            logger.warning(f"Failed to log DocuSign void activity: {log_error}")
        
        flash("✅ DocuSign envelope has been voided. (Mock Mode)", "success")
        logger.info(f"DocuSign envelope voided for contract {contract_id}")
        
        return redirect(url_for("contracts.show", contract_id=contract_id))
        
    except Exception as e:
        logger.error(f"Error voiding DocuSign envelope: {e}")
        flash("An error occurred while voiding the envelope.", "error")
        return redirect(url_for("contracts.show", contract_id=contract_id))


@contracts_bp.route("/admin/populate_docusign_demo", methods=["POST"])
@login_required
def populate_docusign_demo_data():
    """Populate existing contracts with mock DocuSign data for demo purposes"""
    try:
        if not current_user.is_admin:
            flash("Access denied. Admin privileges required.", "error")
            return redirect(url_for("contracts.index"))
        
        result = docusign_service.populate_existing_contracts_with_mock_data()
        
        if result['success']:
            flash(f"✅ Demo data populated: {result['message']}", "success")
        else:
            flash(f"❌ Error populating demo data: {result['error']}", "error")
            
        return redirect(url_for("contracts.index"))
        
    except Exception as e:
        logger.error(f"Error populating DocuSign demo data: {e}")
        flash("An error occurred while populating demo data.", "error")
        return redirect(url_for("contracts.index"))
